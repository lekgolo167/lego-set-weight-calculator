import sys
from math import floor
import requests
import pickle
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook, load_workbook, worksheet
import os


SET_URL = 'https://www.bricklink.com/CatalogItemInv.asp?S='
PART_URL = 'https://www.bricklink.com/v2/catalog/catalogitem.page?P='
FIGURE_URL = 'https://www.bricklink.com/v2/catalog/catalogitem.page?M='
THEME_URL = 'https://brickset.com/sets/theme'
FIGURE_REGEX = r'^[a-zA-Z]{2,6}\d{2,6}$'

inflation_to_2024 = {
	1999: 1.886,
	2000: 1.824,#
	2001: 1.775,#
	2002: 1.746,#
	2003: 1.707,#
	2004: 1.663,#
	2005: 1.609,#
	2006: 1.558,#
	2007: 1.515,#
	2008: 1.459,#
	2009: 1.464,#
	2010: 1.441,
	2011: 1.397,
	2012: 1.368,
	2013: 1.349,
	2014: 1.327,
	2015: 1.326,
	2016: 1.309,
	2017: 1.282,
	2018: 1.251,
	2019: 1.229,
	2020: 1.214,
	2021: 1.159,
	2022: 1.074,
	2023: 1.031,
	2024: 1.000
}

class LegoSet:
	def __init__(self, year:int, name:str, set_id:int, msrp:float, pieces:int):
		self.year = year
		self.name = name
		self.set_id = set_id
		self.msrp = msrp
		self.price = round(inflation_to_2024[year] * msrp, 2)
		self.pieces = pieces
		self.weight = 0.0

def get_theme_years(theme:str) -> list:
	resp = requests.get(url=f'{THEME_URL}-{theme}', headers={'User-Agent': theme})
	years = []
	if resp.status_code == 200:
		try:
			text = resp.text
			m = re.findall(r"year-(\d\d\d\d)'>(\d\d\d\d) ", text)
			if m:
				for year_match in m:
					year = int(year_match[0])
					years.append(year)
				years.sort()
				return years
			else:
				raise RuntimeError(f'Failed to find years for theme {theme}')
		except Exception as e:
			print(f'Failed to get theme ({theme}). Is it spelt correctly (as shown in the URL) and found on brickset.com?')
			print(e)
			sys.exit(1)
	else:
		print(f'Failed to fetch ({theme}) years html\nStatus code: {resp.status_code}')
		sys.exit(1)

def get_sets(theme:str, year:int) -> list:
	set_list = []
	text = ''
	path = f'themes/{theme}/{year}.html'
	exists = os.path.exists(path)
	if exists:
		with open(path, 'r', encoding='utf-8') as f:
			text = f.read()
	else:
		print(f'{theme}:{year} not cached.')
		cookies = {'setsPageLength':'200'}
		resp = requests.get(url=f'{THEME_URL}-{theme}/year-{year}', cookies=cookies)
		text = resp.text
		with open(f'themes/{theme}/{year}.html', 'w', encoding='utf-8') as f:
			f.write(text)
		if resp.status_code == 200:
			text = resp.text
		else:
			print(f'Status code {resp.status_code}')
			return set_list
	try:
		soup = BeautifulSoup(text, 'html.parser')
		sets = soup.find_all(attrs={'class': 'set'})
		for lset in sets:
			try:
				set_info = lset.find(attrs={'class': 'highslide-caption'})
				set_name = set_info.find('h1').get_text()
				set_id = int(set_info.find('a').get_text().replace('-1', ''))
				set_info = lset.find('div', {'class': 'col'})
				piece_count = int(set_info.find('dt', string='Pieces').find_next_sibling().find('a').get_text())
				price_list = set_info.find('dt', string='RRP').find_next_sibling().get_text()
				for p in re.split(',|\|', price_list):
					if '$' in p:
						price = float(p.replace('$', ''))
						break
				else:
					print('NO PRICE FOUND :(')
					price = -1
				#print(f'{set_id}:{set_name}\nContains ({piece_count}) for {price}')
				set_list.append(LegoSet(year, set_name, set_id, price, piece_count))
			except Exception as ee:
				pass
				#print(ee)
	except Exception as e:
		print(e)
	
	return set_list

def is_minifugre(part_id:str) -> bool:
	match = re.search(FIGURE_REGEX, part_id)
	return match is not None

def get_minifigure_weight(fig_id:str) -> float:
	print(f'Fetching Minifigure ({fig_id}).......')
	resp = requests.get(url=f'{FIGURE_URL}{fig_id}', headers={'User-Agent': fig_id})
	if resp.status_code == 200:
		try:
			text = resp.text
			soup = BeautifulSoup(text, 'html.parser')
			weights = soup.find_all(attrs={'id': 'item-weight-info'})
			weight = weights[0]
			return float(weight.get_text().replace('g', ''))
		except Exception as e:
			print(f'Failed to get Minifigure weight for ({fig_id})')
	else:
		print(f'Failed to fetch Minifigure ({fig_id}) info\nStatus code: {resp.status_code}')
	
	return 0.0

def get_part_weight(part_id:str) -> int:
	if is_minifugre(part_id):
		return get_minifigure_weight(part_id)
	print(f'Fetching ({part_id}).......')
	resp = requests.get(url=f'{PART_URL}{part_id}', headers={'User-Agent': part_id})
	if resp.status_code == 200:
		try:
			text = resp.text
			soup = BeautifulSoup(text, 'html.parser')
			weights = soup.find_all(attrs={'id': 'item-weight-info'})
			weight = weights[0]
			return float(weight.get_text().replace('g', ''))
		except Exception as e:
			print(f'Failed to get part ({part_id})')
	else:
		print(f'Failed to get ({part_id})\nStatus code: {resp.status_code}')
	
	return 0.0

def get_set_weight(part_weights:dict, set_parts:dict) -> float:
	set_weight = 0.0
	for set_part_id, quantity in set_parts.items():
		if 'pb' in set_part_id:
			continue
		part_weight = part_weights.get(set_part_id, None)
		if part_weight is None:
			part_weight = get_part_weight(set_part_id)
			part_weights[set_part_id] = part_weight
		#set_weight += floor(part_weight * quantity + 0.3)
		set_weight += part_weight * quantity

	return set_weight

def parse_set(set_id:int) -> dict:
	set_parts = {}
	total_q = 0
	text = ''
	path = f'sets/{set_id}.html'
	exists = os.path.exists(path)
	if exists:
		with open(path, 'r', encoding='utf-8') as f:
			text = f.read()
	else:
		print(f'{set_id} not cached, fetching...', end='')
		resp = requests.get(url=f'{SET_URL}{set_id}-1', headers={'User-Agent': str(set_id)})
		text = resp.text
		print('done')
		with open(path, 'w', encoding='utf-8') as f:
			f.write(text)
		if resp.status_code == 200:
			text = resp.text
		else:
			print(f'Status code {resp.status_code}')
			return set_parts

	soup = BeautifulSoup(text, 'html.parser')
	iv_items = soup.find_all(attrs={'class': 'IV_ITEM'})
	try:
		counterparts = soup.find("b", text="Counterparts:")
		assembly_part = counterparts.find_parent(attrs={'bgcolor': '#000000'}).find_next_sibling(attrs={'class': 'IV_ITEM'})
		i = iv_items.index(assembly_part)
		iv_items = iv_items[:i]
	except:
		pass #no counterparts found
	for iv_item in iv_items:
		part_id = iv_item.attrs['class'][0].replace('IV_', '')
		quanity = int(iv_item.find_all('td')[1].get_text().replace('\xa0', ''))
		if 'stk' in part_id:
			continue
		total_q += quanity
		p = set_parts.get(part_id, None)
		if p is None:
			set_parts[part_id] = quanity
		else:
			set_parts[part_id] += quanity

	print(f'Set {set_id}, has ({total_q}) parts with ({len(set_parts)}) unique.')
	return set_parts

def load_or_create_workbook(file_path:str) -> Workbook:
	if os.path.exists(file_path):
		print(f'Loading existing workbook: {file_path}')
		return load_workbook(file_path)
	print(f'Creating new workbook: {file_path}')
	workbook = Workbook()
	workbook.save(file_path)
	return workbook

def get_or_create_sheet(workbook:Workbook, sheet_name:str) -> worksheet:
	if sheet_name in workbook.sheetnames:
		print(f'Sheet "{sheet_name}" exists. Loading...')
		return workbook[sheet_name]
	print(f'Sheet "{sheet_name}" does not exist. Creating it.')
	workbook.create_sheet(title=sheet_name)
	return workbook[sheet_name]

def fillout_workbook(filename:str, sets_year:list, theme:str) -> None:
	wb = load_or_create_workbook(filename)
	theme_sheet = get_or_create_sheet(wb, theme)
	headers = ['Year', 'Set ID', 'Name', 'MSRP', 'Price', 'Pieces', 'Grams', 'per/gram', 'per/piece']
	
	rows = 1
	years = []
	for sets in sets_year:
		if len(sets) == 0:
			continue
		for i, header in enumerate(headers):
			theme_sheet.cell(row=rows, column=i+1, value=header)
		rows+=1
		years.append(sets[0].year)
		for s in sets:
			theme_sheet.cell(row=rows, column=1, value=s.year)
			theme_sheet.cell(row=rows, column=2, value=s.set_id)
			theme_sheet.cell(row=rows, column=3, value=s.name)
			theme_sheet.cell(row=rows, column=4, value=s.msrp)
			theme_sheet.cell(row=rows, column=5, value=s.price)
			theme_sheet.cell(row=rows, column=6, value=s.pieces)
			theme_sheet.cell(row=rows, column=7, value=s.weight)
			if s.weight > 0.0:
				theme_sheet.cell(row=rows, column=8, value=round(s.price/s.weight*100, 1))
			else:
				theme_sheet.cell(row=rows, column=8, value=0)
			theme_sheet.cell(row=rows, column=9, value=round(s.price/s.pieces*100, 1))
			rows += 1
		rows += 3
	total_r = rows + 1

	headers = ['Year', 'Avg per/gram', 'Avg per/piece']
	for i, header in enumerate(headers):
		theme_sheet.cell(row=1, column=11+i, value=header)

	rows = 2
	for year in years:
		ppg = f'=ROUND(AVERAGEIFS(H2:H{total_r},H2:H{total_r}, ">0", A2:A{total_r}, {year}),1)'
		ppp = f'=ROUND(AVERAGEIFS(I2:I{total_r},I2:I{total_r}, ">0", A2:A{total_r}, {year}),1)'
		theme_sheet.cell(row=rows, column=11, value=year)
		theme_sheet.cell(row=rows, column=12, value=ppg)
		theme_sheet.cell(row=rows, column=13, value=ppp)
		rows += 1

	wb.save(filename=filename)
	wb.close()
	print('Workbook saved')

def create_directories(theme:str) -> None:
	directories = [
		'./parts',
		'./sets'
		'./themes',
		f'./themes/{theme}'
	]
	for directory in directories:
		if not os.path.exists(directory):
			os.makedirs(directory)
			print(f'Created directory: {directory}')

if __name__ == '__main__':
	all_sets = []
	theme = 'alpha-team'
	create_directories(theme)
	cached_parts_path = './parts/parts.pkl'
	if os.path.exists(cached_parts_path):
		part_weights = pickle.load(open(cached_parts_path, 'rb'))
	else:
		part_weights = {}

	years = get_theme_years(theme)
	for year in years:
		print(f'Getting {theme} year {year}')
		s = get_sets(theme, year)
		for sw in s:
			try:
				set_parts = parse_set(sw.set_id)
				sw.weight = round(get_set_weight(part_weights, set_parts), 1)
				print(f'Set {sw.set_id}:{sw.name} weight = {sw.weight}g')
			except Exception as e:
				print(e)
		all_sets.append(s)
		print(f'{year} had {len(s)} sets')
	try:
		fillout_workbook('Lego_sets.xlsx', all_sets, theme)
	except Exception as e:
		print(e)
	print(f'Total cached parts: {len(part_weights)}')
	pickle.dump(part_weights, open('parts/parts.pkl', 'wb'))